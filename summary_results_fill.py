from __future__ import annotations

from datetime import datetime, timedelta, timezone
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import config


GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
TEAM_SUMMARY_SHEET = config.SHEET_TEAM_SUMMARY


def parse_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def infer_side(avg: float, cb: float) -> str:
    if avg < cb:
        return "under"
    if avg > cb:
        return "over"
    return "push"


def normalize_matchup(value) -> str:
    if not isinstance(value, str):
        return ""
    return value.strip().lower()


def fetch_events_lookup(timeout: int) -> dict[str, dict]:
    now_utc = datetime.now(timezone.utc)
    start = datetime(now_utc.year, now_utc.month, now_utc.day, tzinfo=timezone.utc) - timedelta(days=3)
    end = datetime(now_utc.year, now_utc.month, now_utc.day, tzinfo=timezone.utc) + timedelta(days=3)
    params = {"startOfDay": int(start.timestamp()), "endOfDay": int(end.timestamp())}
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json, text/plain, */*",
        "Referer": config.BASE_SITE + "/",
    }
    try:
        resp = requests.get(
            f"{config.BASE_SITE}/api/event/by-date",
            params=params,
            headers=headers,
            timeout=timeout,
        )
        resp.raise_for_status()
    except requests.RequestException:
        return {}

    payload = resp.json()
    rows = payload.get("data", []) if isinstance(payload, dict) else []
    if not isinstance(rows, list):
        return {}

    lookup = {}
    for game in rows:
        if not isinstance(game, dict):
            continue
        home = game.get("homeTeam") or {}
        away = game.get("awayTeam") or {}
        events = game.get("events") or {}
        matchup = f"{home.get('name', '')} - {away.get('name', '')}"
        key = normalize_matchup(matchup)
        if not key:
            continue
        event_id = events.get("id")
        home_team_id = home.get("id")
        if event_id is None or home_team_id is None:
            continue
        lookup[key] = {
            "event_id": int(event_id),
            "home_team_id": int(home_team_id),
            "status": str(events.get("status") or "").strip().lower(),
        }
    return lookup


def fetch_event(match_id: int, timeout: int) -> dict | None:
    url = f"{config.BASE_SITE}/api/event/{match_id}"
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json, text/plain, */*",
        "Referer": config.BASE_SITE + "/",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=timeout)
        resp.raise_for_status()
    except requests.RequestException:
        return None
    payload = resp.json()
    return payload if isinstance(payload, dict) else None


def fetch_team_history(team_id: int, timeout: int) -> list[dict]:
    url = f"{config.BASE_SITE}/api/team/{team_id}/performance"
    params = {
        "limit": config.GAME_STATS_HISTORY_LIMIT,
        "location": "all",
        "eventHalf": "ALL",
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json, text/plain, */*",
        "Referer": config.BASE_SITE + "/",
    }
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=timeout)
        resp.raise_for_status()
    except requests.RequestException:
        return []
    payload = resp.json()
    rows = payload.get("data", []) if isinstance(payload, dict) else []
    return rows if isinstance(rows, list) else []


def find_match_row(history_rows: list[dict], event_id: int) -> dict | None:
    for row in history_rows:
        if not isinstance(row, dict):
            continue
        event = row.get("event") or {}
        if event.get("id") == event_id:
            return row
    return None


def compute_actual_stat_value(match_row: dict | None, stat_label: str, side: str = "total") -> float | None:
    if not match_row:
        return None
    stat_key = config.STATSHUB_STAT_KEYS.get(stat_label)
    if not stat_key:
        return None
    stats = match_row.get("statistics") or {}
    opp_stats = match_row.get("opponentStatistics") or {}
    a = parse_float(stats.get(stat_key))
    b = parse_float(opp_stats.get(stat_key))
    if side == "1":
        return a
    if side == "2":
        return b
    if a is None or b is None:
        return None
    return a + b


def evaluate(side: str, cb_value: float, actual_total: float | None) -> tuple[str, PatternFill]:
    if actual_total is None:
        return "N/A", YELLOW_FILL
    if side == "under":
        return ("GREEN", GREEN_FILL) if actual_total < cb_value else ("RED", RED_FILL)
    if side == "over":
        return ("GREEN", GREEN_FILL) if actual_total > cb_value else ("RED", RED_FILL)
    return ("GREEN", GREEN_FILL) if actual_total == cb_value else ("RED", RED_FILL)


def main():
    input_path: Path = config.EXCEL_PATH

    if not input_path.exists():
        print(f"Skipping summary_results_fill.py: input workbook not found: {input_path}")
        return

    wb = load_workbook(input_path)
    if config.SHEET_SUMMARY not in wb.sheetnames:
        raise SystemExit(f"'{config.SHEET_SUMMARY}' sheet not found in {input_path}")

    # matchup -> match_id from full sheet
    full_df = pd.read_excel(input_path, sheet_name="full")
    summary_df = pd.read_excel(input_path, sheet_name=config.SHEET_SUMMARY)
    team_summary_df = (
        pd.read_excel(input_path, sheet_name=TEAM_SUMMARY_SHEET)
        if TEAM_SUMMARY_SHEET in wb.sheetnames
        else pd.DataFrame()
    )
    matchup_to_match_id = {}
    if "matchup" in full_df.columns and "match_id" in full_df.columns:
        for _, item in full_df.iterrows():
            key = normalize_matchup(item.get("matchup"))
            if not key or key in matchup_to_match_id:
                continue
            match_id = item.get("match_id")
            try:
                matchup_to_match_id[key] = int(match_id)
            except (TypeError, ValueError):
                continue
    ws_summary = wb[config.SHEET_SUMMARY]
    ws_team_summary = wb[TEAM_SUMMARY_SHEET] if TEAM_SUMMARY_SHEET in wb.sheetnames else None
    by_date_lookup = fetch_events_lookup(config.HTTP_TIMEOUT_SECONDS)

    def get_headers(ws) -> dict[str, int]:
        out = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if isinstance(value, str) and value.strip():
                out[value.strip()] = col
        return out

    def ensure_headers(ws, names: list[str]) -> dict[str, int]:
        headers = get_headers(ws)
        next_col = ws.max_column + 1 if ws.max_column > 0 else 1
        if not headers and next_col == 1:
            next_col = 1
        for name in names:
            if name in headers:
                continue
            ws.cell(row=1, column=next_col, value=name)
            headers[name] = next_col
            next_col += 1
        return headers

    def build_row_index(ws, matchup_col: int | None) -> dict[str, int]:
        out = {}
        if matchup_col is None:
            return out
        for r in range(2, ws.max_row + 1):
            key = normalize_matchup(ws.cell(row=r, column=matchup_col).value)
            if key and key not in out:
                out[key] = r
        return out

    summary_headers = get_headers(ws_summary)
    team_summary_headers = get_headers(ws_team_summary) if ws_team_summary is not None else {}
    if "matchup" not in summary_headers:
        raise SystemExit("'matchup' column not found in summary sheet.")

    results_sheet_name = "results"
    ws_results = wb[results_sheet_name] if results_sheet_name in wb.sheetnames else wb.create_sheet(results_sheet_name)
    result_header_names = [
        f"result_{stat}"
        for stat in config.STAT_ORDER
        if f"average_{stat}" in summary_headers and f"cb_{stat}" in summary_headers
    ]
    team_result_header_names = [
        f"result_{stat}_{side}"
        for stat in config.STAT_ORDER
        for side in ("1", "2")
        if f"average_{stat}_{side}" in team_summary_headers and f"cb_{stat}_{side}" in team_summary_headers
    ]
    merged_header_names = list(summary_headers.keys())
    for name in team_summary_headers.keys():
        if name not in merged_header_names:
            merged_header_names.append(name)
    desired_results_headers = merged_header_names + result_header_names + team_result_header_names
    results_headers = ensure_headers(ws_results, desired_results_headers)
    results_row_by_matchup = build_row_index(ws_results, results_headers.get("matchup"))
    result_cols = {stat: results_headers.get(f"result_{stat}") for stat in config.STAT_ORDER}
    team_result_cols = {
        (stat, side): results_headers.get(f"result_{stat}_{side}")
        for stat in config.STAT_ORDER
        for side in ("1", "2")
    }

    results_table_sheet_name = "results_table"
    ws_results_table = (
        wb[results_table_sheet_name]
        if results_table_sheet_name in wb.sheetnames
        else wb.create_sheet(results_table_sheet_name)
    )
    requested_stats = [
        ("corners", "corners"),
        ("cards", "cards"),
        ("shots_on_target", "shots_on_target"),
        ("total-shots", "total_shots"),
        ("fouls", "fouls"),
        ("saves", "goalkeeper_saves"),
        ("throw_ins", "throw_ins"),
        ("offsides", "offsides"),
    ]
    desired_table_headers = ["matchup"] + [h for h, _ in requested_stats] + [f"{h}_1" for h, _ in requested_stats] + [f"{h}_2" for h, _ in requested_stats]
    table_headers = ensure_headers(ws_results_table, desired_table_headers)
    table_row_by_matchup = build_row_index(ws_results_table, table_headers.get("matchup"))
    table_col_by_stat = {stat_label: table_headers[header_name] for header_name, stat_label in requested_stats}
    team_table_col_by_stat = {
        (stat_label, side): table_headers[f"{header_name}_{side}"]
        for header_name, stat_label in requested_stats
        for side in ("1", "2")
    }

    event_cache: dict[int, dict | None] = {}
    history_cache: dict[int, list[dict]] = {}

    summary_matchup_col = summary_headers["matchup"]
    team_summary_row_by_matchup = {}
    if ws_team_summary is not None and "matchup" in team_summary_headers:
        team_matchup_col = team_summary_headers["matchup"]
        for row in range(2, ws_team_summary.max_row + 1):
            key = normalize_matchup(ws_team_summary.cell(row=row, column=team_matchup_col).value)
            if key and key not in team_summary_row_by_matchup:
                team_summary_row_by_matchup[key] = row

    for row in range(2, ws_summary.max_row + 1):
        matchup = ws_summary.cell(row=row, column=summary_matchup_col).value
        matchup_key = normalize_matchup(matchup)
        if not matchup_key:
            continue
        match_id = matchup_to_match_id.get(matchup_key)
        team_row = team_summary_row_by_matchup.get(matchup_key)

        results_row = results_row_by_matchup.get(matchup_key)
        if results_row is None:
            results_row = ws_results.max_row + 1
            results_row_by_matchup[matchup_key] = results_row

        for col_name, src_col in summary_headers.items():
            dst_col = results_headers.get(col_name)
            if dst_col is None:
                continue
            ws_results.cell(row=results_row, column=dst_col, value=ws_summary.cell(row=row, column=src_col).value)
        if ws_team_summary is not None and team_row is not None:
            for col_name, src_col in team_summary_headers.items():
                dst_col = results_headers.get(col_name)
                if dst_col is None:
                    continue
                ws_results.cell(row=results_row, column=dst_col, value=ws_team_summary.cell(row=team_row, column=src_col).value)

        table_row = table_row_by_matchup.get(matchup_key)
        if table_row is None:
            table_row = ws_results_table.max_row + 1
            table_row_by_matchup[matchup_key] = table_row
        ws_results_table.cell(row=table_row, column=table_headers["matchup"], value=matchup)

        event_data = None
        home_team_id = None
        event_status = None
        if match_id is not None:
            if match_id not in event_cache:
                event_cache[match_id] = fetch_event(match_id, config.HTTP_TIMEOUT_SECONDS)
            event_data = event_cache[match_id]
            if event_data:
                events = event_data.get("events") or {}
                event_status = str(events.get("status") or "").strip().lower()
                home_team = event_data.get("homeTeam") or {}
                home_team_id = home_team.get("id")
                if event_data.get("events", {}).get("id"):
                    match_id = int(event_data["events"]["id"])

        if match_id is None and matchup_key in by_date_lookup:
            fallback = by_date_lookup[matchup_key]
            match_id = fallback["event_id"]
            home_team_id = fallback["home_team_id"]
            event_status = fallback["status"]

        match_row = None
        if match_id is not None and home_team_id is not None:
            home_team_id = int(home_team_id)
            if home_team_id not in history_cache:
                history_cache[home_team_id] = fetch_team_history(home_team_id, config.HTTP_TIMEOUT_SECONDS)
            match_row = find_match_row(history_cache[home_team_id], match_id)

        for stat in config.STAT_ORDER:
            avg_col_name = f"average_{stat}"
            cb_col_name = f"cb_{stat}"
            if avg_col_name not in summary_headers or cb_col_name not in summary_headers:
                continue

            avg_col = results_headers.get(avg_col_name)
            cb_col = results_headers.get(cb_col_name)
            result_col = result_cols.get(stat)
            if result_col is None or avg_col is None or cb_col is None:
                continue

            avg_value = parse_float(ws_summary.cell(row=row, column=summary_headers[avg_col_name]).value)
            cb_value = parse_float(ws_summary.cell(row=row, column=summary_headers[cb_col_name]).value)

            if avg_value is None or cb_value is None:
                ws_results.cell(row=results_row, column=result_col, value="N/A").fill = YELLOW_FILL
                ws_results.cell(row=results_row, column=avg_col).fill = YELLOW_FILL
                ws_results.cell(row=results_row, column=cb_col).fill = YELLOW_FILL
                out_col = table_col_by_stat.get(stat)
                if out_col is not None:
                    ws_results_table.cell(row=table_row, column=out_col, value="N/A").fill = YELLOW_FILL
                continue

            side = infer_side(avg_value, cb_value)
            actual_total = compute_actual_stat_value(match_row, stat, "total")

            if event_status and event_status != "finished":
                result_text = f"{side.upper()} {cb_value:g} | not finished"
                fill = YELLOW_FILL
            else:
                outcome, fill = evaluate(side, cb_value, actual_total)
                actual_text = f"{actual_total:g}" if actual_total is not None else "N/A"
                result_text = f"{side.upper()} {cb_value:g} | actual {actual_text} -> {outcome}"

            ws_results.cell(row=results_row, column=result_col, value=result_text).fill = fill
            ws_results.cell(row=results_row, column=avg_col).fill = fill
            ws_results.cell(row=results_row, column=cb_col).fill = fill
            out_col = table_col_by_stat.get(stat)
            if out_col is not None:
                if actual_total is None:
                    ws_results_table.cell(row=table_row, column=out_col, value="N/A").fill = YELLOW_FILL
                else:
                    ws_results_table.cell(row=table_row, column=out_col, value=actual_total).fill = fill

        if ws_team_summary is None or team_row is None:
            continue

        for stat in config.STAT_ORDER:
            for side in ("1", "2"):
                avg_col_name = f"average_{stat}_{side}"
                cb_col_name = f"cb_{stat}_{side}"
                if avg_col_name not in team_summary_headers or cb_col_name not in team_summary_headers:
                    continue

                avg_col = results_headers.get(avg_col_name)
                cb_col = results_headers.get(cb_col_name)
                result_col = team_result_cols.get((stat, side))
                if result_col is None or avg_col is None or cb_col is None:
                    continue

                avg_value = parse_float(ws_team_summary.cell(row=team_row, column=team_summary_headers[avg_col_name]).value)
                cb_value = parse_float(ws_team_summary.cell(row=team_row, column=team_summary_headers[cb_col_name]).value)

                if avg_value is None or cb_value is None:
                    ws_results.cell(row=results_row, column=result_col, value="N/A").fill = YELLOW_FILL
                    ws_results.cell(row=results_row, column=avg_col).fill = YELLOW_FILL
                    ws_results.cell(row=results_row, column=cb_col).fill = YELLOW_FILL
                    out_col = team_table_col_by_stat.get((stat, side))
                    if out_col is not None:
                        ws_results_table.cell(row=table_row, column=out_col, value="N/A").fill = YELLOW_FILL
                    continue

                side_text = infer_side(avg_value, cb_value)
                actual_value = compute_actual_stat_value(match_row, stat, side)

                if event_status and event_status != "finished":
                    result_text = f"{side_text.upper()} {cb_value:g} | not finished"
                    fill = YELLOW_FILL
                else:
                    outcome, fill = evaluate(side_text, cb_value, actual_value)
                    actual_text = f"{actual_value:g}" if actual_value is not None else "N/A"
                    result_text = f"{side_text.upper()} {cb_value:g} | actual {actual_text} -> {outcome}"

                ws_results.cell(row=results_row, column=result_col, value=result_text).fill = fill
                ws_results.cell(row=results_row, column=avg_col).fill = fill
                ws_results.cell(row=results_row, column=cb_col).fill = fill
                out_col = team_table_col_by_stat.get((stat, side))
                if out_col is not None:
                    if actual_value is None:
                        ws_results_table.cell(row=table_row, column=out_col, value="N/A").fill = YELLOW_FILL
                    else:
                        ws_results_table.cell(row=table_row, column=out_col, value=actual_value).fill = fill

    try:
        wb.save(input_path)
        print(f"Saved: {input_path}")
    except PermissionError:
        raise SystemExit(f"Cannot save {input_path}. Close it in Excel and run again.")


if __name__ == "__main__":
    main()
